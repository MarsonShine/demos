from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from video_analysis_pipeline.exporter import export_review_page, export_workbook
from video_analysis_pipeline.models import OverviewRow, Segment


class ExporterTests(unittest.TestCase):
    def test_export_workbook_populates_overview_and_segment_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "dubbing.result.xlsx"
            export_workbook(
                output_path=output_path,
                rows=[(1, 1, "Dan finds a big box.", "丹发现了一个大箱子。", "00:00:05.333", "00:00:09.166")],
                overview_rows=[
                    OverviewRow(
                        education_stage="小学",
                        subject="[167070462398963715]英语",
                        sequence_no=1,
                        movie_name="Dan's Box",
                        video_title="Dan's Box",
                        muted_video="01.mp4",
                        full_video="02.mp4",
                        background_audio="03.mp3",
                        cover_image="01.jpg",
                        video_description="丹打开神秘盒子，蹦出一个调皮的杰克玩偶!",
                        source="[7]绘本配音",
                    )
                ],
            )

            workbook = load_workbook(output_path)
            self.assertEqual(workbook.worksheets[0].cell(row=1, column=1).value, "学段")
            self.assertEqual(workbook.worksheets[0].cell(row=2, column=4).value, "Dan's Box")
            self.assertEqual(workbook.worksheets[1].cell(row=1, column=1).value, "序号")
            self.assertEqual(workbook.worksheets[1].cell(row=2, column=3).value, "Dan finds a big box.")
            self.assertEqual(workbook.worksheets[1].cell(row=1, column=4).value, "分视频文本（中文）")
            self.assertEqual(workbook.worksheets[1].cell(row=2, column=4).value, "丹发现了一个大箱子。")

    def test_exports_review_page_with_video_and_segments(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "review.html"
            segments = [
                Segment(
                    sequence_no=1,
                    segment_no=1,
                    text="Dan finds a big box.",
                    start_ms=5333,
                    end_ms=9166,
                    text_source="srt",
                    quality_flags=["subtitle_aligned"],
                )
            ]

            export_review_page(
                output_path=output_path,
                video_path="02.mp4",
                segments=segments,
                title="Review",
            )

            html = output_path.read_text(encoding="utf-8")
            self.assertIn('src="02.mp4"', html)
            self.assertIn("Dan finds a big box.", html)
            self.assertIn("segments-data", html)
            self.assertIn("playSegment", html)
            self.assertIn("onlyAnomalies", html)
            self.assertIn("backwardStepMs", html)
            self.assertIn("forwardStepMs", html)
            self.assertIn('value="15"', html)
            self.assertIn("moveStartBackwardButton", html)
            self.assertIn("moveStartForwardButton", html)
            self.assertIn("moveEndBackwardButton", html)
            self.assertIn("moveEndForwardButton", html)
            self.assertIn("adjustSegmentBoundary('start_ms', -1)", html)
            self.assertIn("adjustSegmentBoundary('end_ms', 1)", html)
            self.assertIn("saveAdjustmentsButton", html)
            self.assertNotIn("exportAdjustmentsButton", html)
            self.assertIn("postAdjustmentsToApi", html)
            self.assertIn("/api/review/save", html)
            self.assertIn("let isSavingAdjustments = false;", html)
            self.assertIn("replayButton.disabled = segmentButtonsDisabled;", html)
            self.assertIn("saveAdjustmentsButton.disabled = isSavingAdjustments || collectAdjustments().length === 0;", html)
            self.assertIn("py run_pipeline.py review-server --output-root output", html)

    def test_review_page_only_anomalies_filter_ignores_benign_flags(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "review.html"
            segments = [
                Segment(
                    sequence_no=1,
                    segment_no=1,
                    text="Dan finds a big box.",
                    start_ms=5333,
                    end_ms=9166,
                    text_source="srt",
                    quality_flags=["subtitle_aligned"],
                ),
                Segment(
                    sequence_no=1,
                    segment_no=2,
                    text="Hello, Nishal.",
                    start_ms=10_633,
                    end_ms=12_233,
                    text_source="srt",
                    quality_flags=["subtitle_aligned", "alignment_risk", "subtitle_end_shifted"],
                ),
            ]

            export_review_page(
                output_path=output_path,
                video_path="02.mp4",
                segments=segments,
                title="Review",
            )

            html = output_path.read_text(encoding="utf-8")
            self.assertIn("const benignFlags = new Set(['subtitle_aligned']);", html)
            self.assertIn("const anomalyFlags = getAnomalyFlags(segment);", html)
            self.assertIn("anomalyFlags.length === 0", html)
            self.assertIn("Number(item.dataset.index) === index", html)
            self.assertIn("function formatSignedMilliseconds(delta)", html)
            self.assertIn("localStorage.setItem(storageKey, JSON.stringify(payload));", html)
            self.assertIn("localStorage.removeItem(storageKey);", html)
            self.assertIn("commitSavedAdjustments()", html)
            self.assertIn("const defaultApiBase = 'http://127.0.0.1:8765';", html)
            self.assertIn("当前筛选条件隐藏了选中的分段。", html)


if __name__ == "__main__":
    unittest.main()
