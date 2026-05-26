from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from video_analysis_pipeline.config import load_config
from video_analysis_pipeline.config import SegmentationConfig
from video_analysis_pipeline.models import MediaMetadata, OverviewRow, Segment, SubtitleSpan, TimeRange, TranscriptUtterance, WordTiming
from video_analysis_pipeline.pipeline import (
    ProcessedItem,
    _build_leading_title_segment,
    _estimate_difficulty,
    _normalize_segment_order,
    discover_batch_inputs,
    process_batch,
    process_single_overview,
    process_single_video,
)


class PipelineTests(unittest.TestCase):
    def test_process_single_overview_translates_segments_before_writing_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_dir = Path(tmp_dir) / "output"
            output_dir.mkdir(parents=True)
            workbook_path = output_dir / "dubbing.result.xlsx"

            (output_dir / "manifest.json").write_text(
                json.dumps(
                    {
                        "sequence_no": 1,
                        "source_mp4": str(output_dir / "02.mp4"),
                        "overview": {
                            "education_stage": "小学",
                            "subject": "英语",
                            "sequence_no": 1,
                            "movie_name": "Dan's Box",
                            "video_title": "Dan's Box",
                            "muted_video": "01.mp4",
                            "full_video": "02.mp4",
                            "background_audio": "03.mp3",
                            "cover_image": "01.jpg",
                            "video_description": "旧简介",
                            "difficulty": "",
                            "dialogue_audio": "",
                            "topic": "",
                            "source": "[7]绘本配音",
                        },
                        "outputs": {
                            "workbook": str(workbook_path),
                        },
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            (output_dir / "segments.json").write_text(
                json.dumps(
                    {
                        "sequence_no": 1,
                        "segments": [
                            Segment(
                                sequence_no=1,
                                segment_no=1,
                                text="Dan finds a big box.",
                                start_ms=0,
                                end_ms=1_000,
                            ).to_json()
                        ],
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            (output_dir / "02.mp4").write_bytes(b"fake")

            config = load_config(Path("pipeline_config.json"))

            def fake_translate(*, segments: list[Segment], config: object) -> list[Segment]:
                for segment in segments:
                    segment.translated_text = "丹发现了一个大箱子。"
                return segments

            with patch("video_analysis_pipeline.pipeline.generate_video_summary", return_value="新的简介"), patch(
                "video_analysis_pipeline.pipeline.translate_segments_for_education",
                side_effect=fake_translate,
            ) as translate_mock:
                process_single_overview(
                    output_dir=output_dir,
                    config=config,
                    template_path=None,
                    workbook_output=workbook_path,
                )

            translate_mock.assert_called_once()
            workbook = load_workbook(workbook_path)
            self.assertEqual(workbook.worksheets[1].cell(row=2, column=4).value, "丹发现了一个大箱子。")
            updated_segments_payload = json.loads((output_dir / "segments.json").read_text(encoding="utf-8"))
            self.assertEqual(updated_segments_payload["segments"][0]["translated_text"], "丹发现了一个大箱子。")

    def test_discover_batch_inputs_recurses_and_keeps_relative_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_root = Path(tmp_dir) / "input"
            first_job = input_root / "1"
            second_job = input_root / "nested" / "10"
            first_job.mkdir(parents=True)
            second_job.mkdir(parents=True)
            (first_job / "02.mp4").write_bytes(b"fake")
            (first_job / "02.srt").write_text("", encoding="utf-8")
            (first_job / "02.mp3").write_bytes(b"bgm")
            (second_job / "lesson.mp4").write_bytes(b"fake")
            (second_job / "lesson.srt").write_text("", encoding="utf-8")

            items = discover_batch_inputs(input_root)

            self.assertEqual(len(items), 2)
            self.assertEqual(items[0].source_mp4.relative_to(input_root), Path("1") / "02.mp4")
            self.assertEqual(items[0].source_srt.relative_to(input_root), Path("1") / "02.srt")
            self.assertEqual(items[0].source_mp3.relative_to(input_root), Path("1") / "02.mp3")
            self.assertEqual(items[0].relative_dir, Path("1"))
            self.assertEqual(items[1].source_mp4.relative_to(input_root), Path("nested") / "10" / "lesson.mp4")
            self.assertEqual(items[1].source_srt.relative_to(input_root), Path("nested") / "10" / "lesson.srt")
            self.assertIsNone(items[1].source_mp3)
            self.assertEqual(items[1].relative_dir, Path("nested") / "10")

    def test_discover_batch_inputs_raises_when_folder_does_not_have_exact_pair(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_root = Path(tmp_dir) / "input"
            invalid_job = input_root / "broken"
            invalid_job.mkdir(parents=True)
            (invalid_job / "02.mp4").write_bytes(b"fake")

            with self.assertRaisesRegex(RuntimeError, "exactly one MP4, one SRT, and at most one MP3"):
                discover_batch_inputs(input_root)

    def test_discover_batch_inputs_raises_when_folder_has_multiple_mp3_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_root = Path(tmp_dir) / "input"
            invalid_job = input_root / "lesson"
            invalid_job.mkdir(parents=True)
            (invalid_job / "02.mp4").write_bytes(b"fake")
            (invalid_job / "02.srt").write_text("", encoding="utf-8")
            (invalid_job / "a.mp3").write_bytes(b"one")
            (invalid_job / "b.mp3").write_bytes(b"two")

            with self.assertRaisesRegex(RuntimeError, "at most one MP3"):
                discover_batch_inputs(input_root)

    def test_process_batch_mirrors_relative_output_paths(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_root = Path(tmp_dir) / "input"
            output_root = Path(tmp_dir) / "output"
            nested_job = input_root / "season1" / "episode2"
            nested_job.mkdir(parents=True)
            source_mp4 = nested_job / "clip.mp4"
            source_srt = nested_job / "clip.srt"
            source_mp3 = nested_job / "clip.mp3"
            source_mp4.write_bytes(b"fake")
            source_srt.write_text("", encoding="utf-8")
            source_mp3.write_bytes(b"bgm")
            calls: list[tuple[Path, Path, Path | None, Path, int]] = []

            def fake_process_single_video(
                source_mp4: Path,
                output_dir: Path,
                sequence_no: int,
                config: object,
                source_srt: Path | None = None,
                source_mp3: Path | None = None,
                template_path: Path | None = None,
                workbook_output: Path | None = None,
                transcriber: object | None = None,
                progress_callback: object | None = None,
                generate_overview: bool = True,
            ) -> ProcessedItem:
                assert source_srt is not None
                calls.append((source_mp4, source_srt, source_mp3, output_dir, sequence_no))
                return ProcessedItem(
                    sequence_no=sequence_no,
                    source_mp4=source_mp4,
                    output_dir=output_dir,
                    workbook_path=None,
                    review_page_path=None,
                    segments=[],
                )

            with patch("video_analysis_pipeline.pipeline.process_single_video", side_effect=fake_process_single_video), patch(
                "video_analysis_pipeline.pipeline.write_json"
            ):
                results = process_batch(
                    input_root=input_root,
                    output_root=output_root,
                    source_name=None,
                    srt_name=None,
                    config=object(),
                    template_path=None,
                    workbook_output=None,
                )

            self.assertEqual(len(results), 1)
            self.assertEqual(len(calls), 1)
            self.assertEqual(calls[0][0], source_mp4)
            self.assertEqual(calls[0][1], source_srt)
            self.assertEqual(calls[0][2], source_mp3)
            self.assertEqual(calls[0][3], output_root / "season1" / "episode2")
            self.assertEqual(calls[0][4], 1)

    def test_process_batch_can_skip_overview_generation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_root = Path(tmp_dir) / "input"
            output_root = Path(tmp_dir) / "output"
            job_dir = input_root / "lesson-01"
            job_dir.mkdir(parents=True)
            (job_dir / "clip.mp4").write_bytes(b"fake")
            (job_dir / "clip.srt").write_text("", encoding="utf-8")
            received_generate_overview: list[bool] = []

            def fake_process_single_video(
                source_mp4: Path,
                output_dir: Path,
                sequence_no: int,
                config: object,
                source_srt: Path | None = None,
                source_mp3: Path | None = None,
                template_path: Path | None = None,
                workbook_output: Path | None = None,
                transcriber: object | None = None,
                progress_callback: object | None = None,
                generate_overview: bool = True,
            ) -> ProcessedItem:
                received_generate_overview.append(generate_overview)
                return ProcessedItem(
                    sequence_no=sequence_no,
                    source_mp4=source_mp4,
                    output_dir=output_dir,
                    workbook_path=None,
                    review_page_path=None,
                    segments=[],
                    overview_row=None,
                )

            with patch("video_analysis_pipeline.pipeline.process_single_video", side_effect=fake_process_single_video), patch(
                "video_analysis_pipeline.pipeline.export_workbook"
            ) as export_workbook_mock:
                process_batch(
                    input_root=input_root,
                    output_root=output_root,
                    source_name=None,
                    srt_name=None,
                    config=object(),
                    template_path=None,
                    workbook_output=output_root / "dubbing.result.xlsx",
                    generate_overview=False,
                )

            self.assertEqual(received_generate_overview, [False])
            export_workbook_mock.assert_not_called()

    def test_process_batch_mod_outputs_sequence_directories_and_removes_intermediate_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_root = Path(tmp_dir) / "input"
            output_root = Path(tmp_dir) / "output"
            job_dir = input_root / "story-pack" / "A Fun Day Out"
            job_dir.mkdir(parents=True)
            source_mp4 = job_dir / "clip.mp4"
            source_srt = job_dir / "clip.srt"
            source_mp4.write_bytes(b"fake")
            source_srt.write_text("", encoding="utf-8")

            def fake_process_single_video(
                source_mp4: Path,
                output_dir: Path,
                sequence_no: int,
                config: object,
                source_srt: Path | None = None,
                source_mp3: Path | None = None,
                template_path: Path | None = None,
                workbook_output: Path | None = None,
                transcriber: object | None = None,
                progress_callback: object | None = None,
                generate_overview: bool = True,
            ) -> ProcessedItem:
                output_dir.mkdir(parents=True, exist_ok=True)
                for file_name in [
                    "01.jpg",
                    "01.mp4",
                    "02.mp4",
                    "03.mp3",
                    "manifest.json",
                    "progress.json",
                    "review.html",
                    "segments.csv",
                    "segments.json",
                    "subtitle_spans.json",
                ]:
                    (output_dir / file_name).write_text("artifact", encoding="utf-8")
                return ProcessedItem(
                    sequence_no=sequence_no,
                    source_mp4=source_mp4,
                    output_dir=output_dir,
                    workbook_path=None,
                    review_page_path=output_dir / "review.html",
                    segments=[
                        Segment(
                            sequence_no=sequence_no,
                            segment_no=1,
                            text="A fun day out.",
                            start_ms=0,
                            end_ms=1_000,
                        )
                    ],
                    overview_row=OverviewRow(
                        education_stage="小学",
                        subject="英语",
                        sequence_no=sequence_no,
                        movie_name="A Fun Day Out",
                        video_title="A Fun Day Out",
                        muted_video="01.mp4",
                        full_video="02.mp4",
                        background_audio="03.mp3",
                        cover_image="01.jpg",
                        video_description="A short story.",
                        difficulty="简单",
                        source="绘本配音",
                    ),
                )

            with patch("video_analysis_pipeline.pipeline.process_single_video", side_effect=fake_process_single_video):
                results = process_batch(
                    input_root=input_root,
                    output_root=output_root,
                    source_name=None,
                    srt_name=None,
                    config=object(),
                    template_path=None,
                    workbook_output=None,
                    final_output="mod",
                )

            self.assertEqual(len(results), 1)
            item_dir = output_root / "dubbing" / "1"
            self.assertEqual(results[0].output_dir, item_dir)
            self.assertTrue((output_root / "movie_dubbing.xlsx").exists())
            self.assertTrue((item_dir / "01.jpg").exists())
            self.assertTrue((item_dir / "01.mp4").exists())
            self.assertTrue((item_dir / "02.mp4").exists())
            self.assertTrue((item_dir / "03.mp3").exists())
            self.assertFalse((output_root / "batch_summary.json").exists())
            self.assertFalse((output_root / "batch_progress.json").exists())
            self.assertFalse((item_dir / "manifest.json").exists())
            self.assertFalse((item_dir / "progress.json").exists())
            self.assertFalse((item_dir / "review.html").exists())
            self.assertFalse((item_dir / "segments.csv").exists())
            self.assertFalse((item_dir / "segments.json").exists())
            self.assertFalse((item_dir / "subtitle_spans.json").exists())

            workbook = load_workbook(output_root / "movie_dubbing.xlsx")
            self.assertEqual(workbook.worksheets[0].cell(row=2, column=3).value, 1)
            self.assertEqual(workbook.worksheets[0].cell(row=2, column=4).value, "A Fun Day Out")

    def test_process_batch_resume_skips_completed_items(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_root = Path(tmp_dir) / "input"
            output_root = Path(tmp_dir) / "output"
            first_job = input_root / "lesson-01"
            second_job = input_root / "lesson-02"
            first_job.mkdir(parents=True)
            second_job.mkdir(parents=True)
            (first_job / "clip.mp4").write_bytes(b"fake")
            (first_job / "clip.srt").write_text("", encoding="utf-8")
            (second_job / "clip.mp4").write_bytes(b"fake")
            (second_job / "clip.srt").write_text("", encoding="utf-8")

            resumed_output_dir = output_root / "lesson-01"
            resumed_output_dir.mkdir(parents=True)
            (resumed_output_dir / "manifest.json").write_text(
                json.dumps(
                    {
                        "sequence_no": 1,
                        "source_mp4": str(resumed_output_dir / "02.mp4"),
                        "outputs": {
                            "review_html": None,
                            "workbook": None,
                        },
                        "timings_seconds": {"copy-source-video": 1.25},
                    }
                ),
                encoding="utf-8",
            )
            (resumed_output_dir / "segments.json").write_text(
                json.dumps(
                    {
                        "sequence_no": 1,
                        "source_mp4": str(resumed_output_dir / "02.mp4"),
                        "segments": [],
                    }
                ),
                encoding="utf-8",
            )
            (resumed_output_dir / "subtitle_spans.json").write_text(
                json.dumps({"subtitle_spans": []}),
                encoding="utf-8",
            )
            (output_root / "batch_progress.json").write_text(
                json.dumps(
                    {
                        "items": [
                            {
                                "output_dir": str(resumed_output_dir),
                                "status": "completed",
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )

            processed_output_dirs: list[Path] = []

            def fake_process_single_video(
                source_mp4: Path,
                output_dir: Path,
                sequence_no: int,
                config: object,
                source_srt: Path | None = None,
                source_mp3: Path | None = None,
                template_path: Path | None = None,
                workbook_output: Path | None = None,
                transcriber: object | None = None,
                progress_callback: object | None = None,
                generate_overview: bool = True,
            ) -> ProcessedItem:
                processed_output_dirs.append(output_dir)
                return ProcessedItem(
                    sequence_no=sequence_no,
                    source_mp4=source_mp4,
                    output_dir=output_dir,
                    workbook_path=None,
                    review_page_path=None,
                    segments=[],
                )

            with patch("video_analysis_pipeline.pipeline.process_single_video", side_effect=fake_process_single_video):
                results = process_batch(
                    input_root=input_root,
                    output_root=output_root,
                    source_name=None,
                    srt_name=None,
                    config=object(),
                    template_path=None,
                    workbook_output=None,
                    generate_overview=False,
                    resume=True,
                )

            self.assertEqual(processed_output_dirs, [output_root / "lesson-02"])
            self.assertEqual([item.output_dir for item in results], [output_root / "lesson-01", output_root / "lesson-02"])
            self.assertEqual(results[0].timings, {"copy-source-video": 1.25})

    def test_process_batch_resume_recovers_completed_items_from_existing_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_root = Path(tmp_dir) / "input"
            output_root = Path(tmp_dir) / "output"
            first_job = input_root / "lesson-01"
            second_job = input_root / "lesson-02"
            first_job.mkdir(parents=True)
            second_job.mkdir(parents=True)
            (first_job / "clip.mp4").write_bytes(b"fake")
            (first_job / "clip.srt").write_text("", encoding="utf-8")
            (second_job / "clip.mp4").write_bytes(b"fake")
            (second_job / "clip.srt").write_text("", encoding="utf-8")

            resumed_output_dir = output_root / "dubbing" / "1"
            resumed_output_dir.mkdir(parents=True)
            (resumed_output_dir / "manifest.json").write_text(
                json.dumps(
                    {
                        "sequence_no": 1,
                        "source_mp4": str(resumed_output_dir / "02.mp4"),
                        "outputs": {
                            "review_html": None,
                            "workbook": None,
                        },
                        "timings_seconds": {"copy-source-video": 1.25},
                    }
                ),
                encoding="utf-8",
            )
            (resumed_output_dir / "segments.json").write_text(
                json.dumps(
                    {
                        "sequence_no": 1,
                        "source_mp4": str(resumed_output_dir / "02.mp4"),
                        "segments": [],
                    }
                ),
                encoding="utf-8",
            )
            (resumed_output_dir / "subtitle_spans.json").write_text(
                json.dumps({"subtitle_spans": []}),
                encoding="utf-8",
            )
            (output_root / "batch_progress.json").write_text(
                json.dumps(
                    {
                        "items": [
                            {
                                "output_dir": str(resumed_output_dir),
                                "status": "running",
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )

            processed_output_dirs: list[Path] = []

            def fake_process_single_video(
                source_mp4: Path,
                output_dir: Path,
                sequence_no: int,
                config: object,
                source_srt: Path | None = None,
                source_mp3: Path | None = None,
                template_path: Path | None = None,
                workbook_output: Path | None = None,
                transcriber: object | None = None,
                progress_callback: object | None = None,
                generate_overview: bool = True,
            ) -> ProcessedItem:
                processed_output_dirs.append(output_dir)
                return ProcessedItem(
                    sequence_no=sequence_no,
                    source_mp4=source_mp4,
                    output_dir=output_dir,
                    workbook_path=None,
                    review_page_path=None,
                    segments=[],
                )

            with patch("video_analysis_pipeline.pipeline.process_single_video", side_effect=fake_process_single_video):
                results = process_batch(
                    input_root=input_root,
                    output_root=output_root,
                    source_name=None,
                    srt_name=None,
                    config=object(),
                    template_path=None,
                    workbook_output=None,
                    generate_overview=False,
                    final_output="mod",
                    resume=True,
                )

            self.assertEqual(processed_output_dirs, [output_root / "dubbing" / "2"])
            self.assertEqual([item.output_dir for item in results], [output_root / "dubbing" / "1", output_root / "dubbing" / "2"])
            self.assertEqual(results[0].timings, {"copy-source-video": 1.25})

    def test_process_batch_failure_persists_completed_items_in_progress(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            input_root = Path(tmp_dir) / "input"
            output_root = Path(tmp_dir) / "output"
            for name in ["lesson-01", "lesson-02", "lesson-03"]:
                job_dir = input_root / name
                job_dir.mkdir(parents=True)
                (job_dir / "clip.mp4").write_bytes(b"fake")
                (job_dir / "clip.srt").write_text("", encoding="utf-8")

            def fake_process_single_video(
                source_mp4: Path,
                output_dir: Path,
                sequence_no: int,
                config: object,
                source_srt: Path | None = None,
                source_mp3: Path | None = None,
                template_path: Path | None = None,
                workbook_output: Path | None = None,
                transcriber: object | None = None,
                progress_callback: object | None = None,
                generate_overview: bool = True,
            ) -> ProcessedItem:
                output_dir.mkdir(parents=True, exist_ok=True)
                if sequence_no < 3:
                    (output_dir / "manifest.json").write_text(
                        json.dumps(
                            {
                                "sequence_no": sequence_no,
                                "source_mp4": str(output_dir / "02.mp4"),
                            }
                        ),
                        encoding="utf-8",
                    )
                    (output_dir / "segments.json").write_text(
                        json.dumps(
                            {
                                "sequence_no": sequence_no,
                                "source_mp4": str(output_dir / "02.mp4"),
                                "segments": [],
                            }
                        ),
                        encoding="utf-8",
                    )
                    return ProcessedItem(
                        sequence_no=sequence_no,
                        source_mp4=source_mp4,
                        output_dir=output_dir,
                        workbook_path=None,
                        review_page_path=None,
                        segments=[],
                    )
                raise RuntimeError("boom on 3")

            with patch("video_analysis_pipeline.pipeline.process_single_video", side_effect=fake_process_single_video):
                with self.assertRaisesRegex(RuntimeError, "boom on 3"):
                    process_batch(
                        input_root=input_root,
                        output_root=output_root,
                        source_name=None,
                        srt_name=None,
                        config=object(),
                        template_path=None,
                        workbook_output=None,
                        generate_overview=False,
                    )

            batch_progress_payload = json.loads((output_root / "batch_progress.json").read_text(encoding="utf-8"))
            self.assertEqual(batch_progress_payload["completed_items"], 2)
            self.assertEqual(
                [(item["sequence_no"], item["status"]) for item in batch_progress_payload["items"]],
                [(1, "completed"), (2, "completed"), (3, "running")],
            )

    def test_process_single_video_can_skip_background_audio_output(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workspace_root = Path(tmp_dir)
            source_mp4 = workspace_root / "clip.mp4"
            source_srt = workspace_root / "clip.srt"
            output_dir = workspace_root / "output"
            source_mp4.write_bytes(b"fake")
            source_srt.write_text("", encoding="utf-8")
            config = load_config(Path(__file__).resolve().parents[1] / "pipeline_config.json")
            config.steps.export_source_video = False
            config.steps.export_cover = False
            config.steps.export_muted_video = False
            config.steps.export_background_audio = False
            config.steps.generate_summary = False
            config.steps.export_workbook = False
            config.steps.export_review_page = False
            config.steps.export_csv = False

            class FakeTranscriber:
                def transcribe(self, audio_path: Path) -> list[TranscriptUtterance]:
                    return []

            def fake_extract_audio_mp3(source_path: Path, output_path: Path) -> Path:
                output_path.write_bytes(b"analysis")
                return output_path

            def fake_probe_media(path: Path) -> MediaMetadata:
                if path.suffix.lower() == ".mp4":
                    return MediaMetadata(
                        path=str(path),
                        duration_ms=1_000,
                        video_streams=1,
                        audio_streams=1,
                        width=1280,
                        height=720,
                        sample_rate=44100,
                        channels=2,
                    )
                return MediaMetadata(
                    path=str(path),
                    duration_ms=1_000,
                    video_streams=0,
                    audio_streams=1,
                    sample_rate=44100,
                    channels=2,
                )

            alignment_summary = {
                "alignment_mode": "asr-only",
                "matched_segments": 0,
                "unmatched_segments": 0,
                "total_subtitle_spans": 0,
            }

            with patch("video_analysis_pipeline.pipeline.extract_audio_mp3", side_effect=fake_extract_audio_mp3), patch(
                "video_analysis_pipeline.pipeline.probe_media",
                side_effect=fake_probe_media,
            ), patch(
                "video_analysis_pipeline.pipeline.detect_silence",
                return_value=([], [TimeRange(start_ms=0, end_ms=1_000)]),
            ), patch(
                "video_analysis_pipeline.pipeline.parse_srt_file",
                return_value=[],
            ), patch(
                "video_analysis_pipeline.pipeline._build_output_segments",
                return_value=([], alignment_summary, []),
            ):
                result = process_single_video(
                    source_mp4=source_mp4,
                    output_dir=output_dir,
                    sequence_no=1,
                    config=config,
                    source_srt=source_srt,
                    transcriber=FakeTranscriber(),
                    generate_overview=False,
                )

            self.assertEqual(result.output_dir, output_dir)
            self.assertFalse((output_dir / "03.mp3").exists())
            manifest_payload = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))
            self.assertIsNone(manifest_payload["audio_mp3"])
            self.assertEqual(manifest_payload["audio_processing"]["output_audio_kind"], "none")
            self.assertIsNone(manifest_payload["media"]["audio"])
            segments_payload = json.loads((output_dir / "segments.json").read_text(encoding="utf-8"))
            self.assertIsNone(segments_payload["audio_mp3"])

    def test_process_single_video_reuses_input_mp3_as_background_audio(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workspace_root = Path(tmp_dir)
            source_mp4 = workspace_root / "clip.mp4"
            source_srt = workspace_root / "clip.srt"
            source_mp3 = workspace_root / "clip.mp3"
            output_dir = workspace_root / "output"
            source_mp4.write_bytes(b"fake")
            source_srt.write_text("", encoding="utf-8")
            source_mp3.write_bytes(b"provided-bgm")
            config = load_config(Path(__file__).resolve().parents[1] / "pipeline_config.json")
            config.steps.export_source_video = False
            config.steps.export_cover = False
            config.steps.export_muted_video = False
            config.steps.export_background_audio = False
            config.steps.generate_summary = False
            config.steps.export_workbook = False
            config.steps.export_review_page = False
            config.steps.export_csv = False

            class FakeTranscriber:
                def transcribe(self, audio_path: Path) -> list[TranscriptUtterance]:
                    return []

            def fake_extract_audio_mp3(source_path: Path, output_path: Path) -> Path:
                output_path.write_bytes(b"analysis")
                return output_path

            def fake_probe_media(path: Path) -> MediaMetadata:
                if path.suffix.lower() == ".mp4":
                    return MediaMetadata(
                        path=str(path),
                        duration_ms=1_000,
                        video_streams=1,
                        audio_streams=1,
                        width=1280,
                        height=720,
                        sample_rate=44100,
                        channels=2,
                    )
                return MediaMetadata(
                    path=str(path),
                    duration_ms=1_000,
                    video_streams=0,
                    audio_streams=1,
                    sample_rate=44100,
                    channels=2,
                )

            alignment_summary = {
                "alignment_mode": "asr-only",
                "matched_segments": 0,
                "unmatched_segments": 0,
                "total_subtitle_spans": 0,
            }

            with patch("video_analysis_pipeline.pipeline.extract_audio_mp3", side_effect=fake_extract_audio_mp3), patch(
                "video_analysis_pipeline.pipeline.extract_background_audio_mp3"
            ) as extract_background_audio_mock, patch(
                "video_analysis_pipeline.pipeline.probe_media",
                side_effect=fake_probe_media,
            ), patch(
                "video_analysis_pipeline.pipeline.detect_silence",
                return_value=([], [TimeRange(start_ms=0, end_ms=1_000)]),
            ), patch(
                "video_analysis_pipeline.pipeline.parse_srt_file",
                return_value=[],
            ), patch(
                "video_analysis_pipeline.pipeline._build_output_segments",
                return_value=([], alignment_summary, []),
            ):
                process_single_video(
                    source_mp4=source_mp4,
                    output_dir=output_dir,
                    sequence_no=1,
                    config=config,
                    source_srt=source_srt,
                    source_mp3=source_mp3,
                    transcriber=FakeTranscriber(),
                    generate_overview=False,
                )

            extract_background_audio_mock.assert_not_called()
            self.assertEqual((output_dir / "03.mp3").read_bytes(), b"provided-bgm")
            manifest_payload = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest_payload["audio_mp3"], str(output_dir / "03.mp3"))
            self.assertEqual(manifest_payload["audio_processing"]["output_audio_kind"], "bgm")
            self.assertEqual(manifest_payload["audio_processing"]["provided_source_path"], str(source_mp3))
            self.assertIsNone(manifest_payload["audio_processing"]["separation_method"])

    def test_process_single_overview_rebuilds_workbook_from_existing_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_dir = Path(tmp_dir) / "07 The Lion And The Mouse"
            output_dir.mkdir(parents=True)
            source_mp4 = output_dir / "02.mp4"
            source_mp4.write_bytes(b"fake")
            (output_dir / "01.mp4").write_bytes(b"fake")
            (output_dir / "03.mp3").write_bytes(b"fake")
            (output_dir / "01.jpg").write_bytes(b"fake")

            segments = [
                Segment(
                    sequence_no=7,
                    segment_no=1,
                    text="The lion roars loudly.",
                    start_ms=1_000,
                    end_ms=2_000,
                    text_source="srt",
                    source_subtitle_index=0,
                )
            ]
            subtitle_spans = [
                SubtitleSpan(
                    text="The lion roars loudly.",
                    normalized_text="the lion roars loudly",
                    start_ms=1_000,
                    end_ms=2_000,
                    confidence=1.0,
                    frame_count=1,
                    source="srt",
                    raw_index=1,
                )
            ]

            (output_dir / "manifest.json").write_text(
                json.dumps(
                    {
                        "sequence_no": 7,
                        "source_mp4": str(source_mp4),
                        "muted_video": str(output_dir / "01.mp4"),
                        "audio_mp3": str(output_dir / "03.mp3"),
                        "cover_image": str(output_dir / "01.jpg"),
                        "outputs": {
                            "review_html": str(output_dir / "review.html"),
                        },
                        "timings_seconds": {"build-segments": 1.25},
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            (output_dir / "segments.json").write_text(
                json.dumps(
                    {
                        "sequence_no": 7,
                        "subtitle": {"spans": [item.to_json() for item in subtitle_spans]},
                        "segments": [item.to_json() for item in segments],
                        "overview": None,
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )

            config = load_config(Path(__file__).resolve().parents[1] / "pipeline_config.json")
            with patch("video_analysis_pipeline.pipeline.generate_video_summary", return_value="狮子和老鼠学会互相帮助。"), patch(
                "video_analysis_pipeline.pipeline.translate_segments_for_education",
                side_effect=lambda *, segments, config: [setattr(segment, "translated_text", f"中文：{segment.text}") or segment for segment in segments],
            ):
                result = process_single_overview(output_dir=output_dir, config=config)

            self.assertEqual(result.sequence_no, 7)
            self.assertIsNotNone(result.workbook_path)
            workbook = load_workbook(result.workbook_path)
            self.assertEqual(workbook.worksheets[0].cell(row=2, column=4).value, "The Lion And The Mouse")
            self.assertEqual(workbook.worksheets[0].cell(row=2, column=10).value, "狮子和老鼠学会互相帮助。")
            self.assertEqual(workbook.worksheets[0].cell(row=2, column=11).value, "1")
            self.assertEqual(workbook.worksheets[1].cell(row=2, column=3).value, "The lion roars loudly.")
            self.assertEqual(workbook.worksheets[1].cell(row=2, column=4).value, "中文：The lion roars loudly.")

            manifest_payload = json.loads((output_dir / "manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest_payload["overview"]["video_title"], "The Lion And The Mouse")
            self.assertEqual(manifest_payload["overview"]["video_description"], "狮子和老鼠学会互相帮助。")
            self.assertEqual(manifest_payload["overview"]["difficulty"], "1")
            self.assertIn("build-segments", manifest_payload["timings_seconds"])
            self.assertIn("generate-video-summary", manifest_payload["timings_seconds"])

            segments_payload = json.loads((output_dir / "segments.json").read_text(encoding="utf-8"))
            self.assertEqual(segments_payload["overview"]["movie_name"], "The Lion And The Mouse")
            self.assertEqual(segments_payload["segments"][0]["translated_text"], "中文：The lion roars loudly.")

    def test_estimate_difficulty_scores_subtitle_complexity_on_1_to_5_scale(self) -> None:
        simple_spans = [
            SubtitleSpan(
                text="The lion roars loudly.",
                normalized_text="the lion roars loudly",
                start_ms=0,
                end_ms=1_000,
                confidence=1.0,
                frame_count=1,
                source="srt",
                raw_index=1,
            )
        ]
        complex_spans = [
            SubtitleSpan(
                text="Although the adventurous children were whispering nervously, they still decided to investigate the mysterious footprints because the strangely illuminated forest seemed unexpectedly welcoming.",
                normalized_text="although the adventurous children were whispering nervously they still decided to investigate the mysterious footprints because the strangely illuminated forest seemed unexpectedly welcoming",
                start_ms=0,
                end_ms=4_000,
                confidence=1.0,
                frame_count=1,
                source="srt",
                raw_index=1,
            ),
            SubtitleSpan(
                text="When the conversation becomes more complicated, the sentences contain subordinate clauses, unusual vocabulary, and longer descriptive phrases.",
                normalized_text="when the conversation becomes more complicated the sentences contain subordinate clauses unusual vocabulary and longer descriptive phrases",
                start_ms=4_000,
                end_ms=8_000,
                confidence=1.0,
                frame_count=1,
                source="srt",
                raw_index=2,
            ),
        ]

        simple_difficulty = _estimate_difficulty(simple_spans, [])
        complex_difficulty = _estimate_difficulty(complex_spans, [])

        self.assertEqual(simple_difficulty, "1")
        self.assertGreaterEqual(int(complex_difficulty), 4)
        self.assertLessEqual(int(complex_difficulty), 5)

    def test_normalize_segment_order_sorts_by_time_and_renumbers(self) -> None:
        segments = [
            Segment(sequence_no=1, segment_no=1, text="Late", start_ms=5_000, end_ms=6_000, source_subtitle_index=1),
            Segment(sequence_no=1, segment_no=2, text="Early", start_ms=1_000, end_ms=2_000, source_subtitle_index=0),
        ]

        _normalize_segment_order(segments)

        self.assertEqual([(segment.segment_no, segment.text, segment.start_ms) for segment in segments], [(1, "Early", 1_000), (2, "Late", 5_000)])

    def test_builds_leading_title_segment_from_asr_before_first_srt(self) -> None:
        subtitle_spans = [
            SubtitleSpan(
                text="Dan finds a big box.",
                normalized_text="dan finds a big box",
                start_ms=5333,
                end_ms=9166,
                confidence=1.0,
                frame_count=1,
                source="srt",
                raw_index=2,
            )
        ]
        utterances = [
            TranscriptUtterance(
                text="Looking for Dragons by Richard Brown and Kate Ruttle.",
                start_ms=0,
                end_ms=5320,
                words=[
                    WordTiming(text="Looking", start_ms=0, end_ms=980, confidence=0.40),
                    WordTiming(text="for", start_ms=980, end_ms=1320, confidence=0.84),
                    WordTiming(text="Dragons", start_ms=1320, end_ms=1600, confidence=0.79),
                    WordTiming(text="by", start_ms=1600, end_ms=3000, confidence=0.65),
                    WordTiming(text="Richard", start_ms=3000, end_ms=3360, confidence=0.99),
                    WordTiming(text="Brown", start_ms=3360, end_ms=3800, confidence=0.97),
                    WordTiming(text="and", start_ms=3800, end_ms=4280, confidence=0.92),
                    WordTiming(text="Kate", start_ms=4280, end_ms=4760, confidence=0.88),
                    WordTiming(text="Ruttle.", start_ms=4760, end_ms=5320, confidence=0.86),
                ],
            ),
            TranscriptUtterance(
                text="Dan finds a big box.",
                start_ms=7040,
                end_ms=9240,
                words=[
                    WordTiming(text="Dan", start_ms=7040, end_ms=7400, confidence=0.99),
                    WordTiming(text="finds", start_ms=7400, end_ms=7800, confidence=0.98),
                    WordTiming(text="a", start_ms=7800, end_ms=7940, confidence=0.94),
                    WordTiming(text="big", start_ms=7940, end_ms=8260, confidence=0.99),
                    WordTiming(text="box.", start_ms=8260, end_ms=9120, confidence=0.96),
                ],
            ),
        ]
        aligned_segments = [
            Segment(
                sequence_no=1,
                segment_no=1,
                text="Dan finds a big box.",
                start_ms=7040,
                end_ms=9120,
                text_source="srt",
                source_subtitle_index=0,
                source_word_range=[9, 13],
                source_utterance_indexes=[1],
            )
        ]

        segment = _build_leading_title_segment(
            sequence_no=1,
            subtitle_spans=subtitle_spans,
            utterances=utterances,
            audio_duration_ms=41_378,
            video_duration_ms=41_533,
            config=SegmentationConfig(),
            aligned_segments=aligned_segments,
        )

        self.assertIsNotNone(segment)
        assert segment is not None
        self.assertEqual(segment.text, "Looking for Dragons by Richard Brown and Kate Ruttle.")
        self.assertEqual(segment.text_source, "asr-title")
        self.assertIn("title_segment_from_asr", segment.quality_flags)
        self.assertEqual(segment.source_word_range, [0, 8])

    def test_title_segment_stops_before_first_aligned_subtitle_words(self) -> None:
        subtitle_spans = [
            SubtitleSpan(
                text="What can we do today?",
                normalized_text="what can we do today",
                start_ms=10880,
                end_ms=13013,
                confidence=1.0,
                frame_count=1,
                source="srt",
                raw_index=1,
            )
        ]
        utterances = [
            TranscriptUtterance(
                text="Looking for Dragons by Richard Brown and Kate Ruttle.",
                start_ms=0,
                end_ms=5320,
                words=[
                    WordTiming(text="Looking", start_ms=0, end_ms=980, confidence=0.40),
                    WordTiming(text="for", start_ms=980, end_ms=1320, confidence=0.84),
                    WordTiming(text="Dragons", start_ms=1320, end_ms=1600, confidence=0.79),
                    WordTiming(text="by", start_ms=1600, end_ms=3000, confidence=0.65),
                    WordTiming(text="Richard", start_ms=3000, end_ms=3360, confidence=0.99),
                    WordTiming(text="Brown", start_ms=3360, end_ms=3800, confidence=0.97),
                    WordTiming(text="and", start_ms=3800, end_ms=4280, confidence=0.92),
                    WordTiming(text="Kate", start_ms=4280, end_ms=4760, confidence=0.88),
                    WordTiming(text="Ruttle.", start_ms=4760, end_ms=5320, confidence=0.86),
                ],
            ),
            TranscriptUtterance(
                text="What can we do today? Look for dragons!",
                start_ms=10080,
                end_ms=16140,
                words=[
                    WordTiming(text="What", start_ms=10080, end_ms=10520, confidence=0.95),
                    WordTiming(text="can", start_ms=10520, end_ms=10820, confidence=0.98),
                    WordTiming(text="we", start_ms=10820, end_ms=10980, confidence=0.98),
                    WordTiming(text="do", start_ms=10980, end_ms=11200, confidence=0.98),
                    WordTiming(text="today?", start_ms=11200, end_ms=13020, confidence=0.92),
                    WordTiming(text="Look", start_ms=13890, end_ms=14510, confidence=0.88),
                ],
            ),
        ]
        aligned_segments = [
            Segment(
                sequence_no=1,
                segment_no=1,
                text="What can we do today?",
                start_ms=10015,
                end_ms=13019,
                text_source="srt",
                source_subtitle_index=0,
                source_word_range=[9, 13],
                source_utterance_indexes=[1],
            )
        ]

        segment = _build_leading_title_segment(
            sequence_no=1,
            subtitle_spans=subtitle_spans,
            utterances=utterances,
            audio_duration_ms=41_378,
            video_duration_ms=41_533,
            config=SegmentationConfig(),
            aligned_segments=aligned_segments,
        )

        self.assertIsNotNone(segment)
        assert segment is not None
        self.assertEqual(segment.text, "Looking for Dragons by Richard Brown and Kate Ruttle.")
        self.assertEqual(segment.source_word_range, [0, 8])
        self.assertLess(segment.end_ms, aligned_segments[0].start_ms)

    def test_title_segment_is_clamped_before_first_aligned_segment_start(self) -> None:
        subtitle_spans = [
            SubtitleSpan(
                text="What can we do today?",
                normalized_text="what can we do today",
                start_ms=10880,
                end_ms=13013,
                confidence=1.0,
                frame_count=1,
                source="srt",
                raw_index=1,
            )
        ]
        utterances = [
            TranscriptUtterance(
                text="Looking for Dragons",
                start_ms=0,
                end_ms=9990,
                words=[
                    WordTiming(text="Looking", start_ms=0, end_ms=3200, confidence=0.90),
                    WordTiming(text="for", start_ms=3200, end_ms=6400, confidence=0.90),
                    WordTiming(text="Dragons", start_ms=6400, end_ms=9990, confidence=0.90),
                ],
            ),
            TranscriptUtterance(
                text="What can we do today?",
                start_ms=10080,
                end_ms=13020,
                words=[
                    WordTiming(text="What", start_ms=10080, end_ms=10520, confidence=0.95),
                    WordTiming(text="can", start_ms=10520, end_ms=10820, confidence=0.98),
                    WordTiming(text="we", start_ms=10820, end_ms=10980, confidence=0.98),
                    WordTiming(text="do", start_ms=10980, end_ms=11200, confidence=0.98),
                    WordTiming(text="today?", start_ms=11200, end_ms=13020, confidence=0.92),
                ],
            ),
        ]
        aligned_segments = [
            Segment(
                sequence_no=1,
                segment_no=1,
                text="What can we do today?",
                start_ms=10015,
                end_ms=13019,
                text_source="srt",
                source_subtitle_index=0,
                source_word_range=[3, 7],
                source_utterance_indexes=[1],
            )
        ]

        segment = _build_leading_title_segment(
            sequence_no=1,
            subtitle_spans=subtitle_spans,
            utterances=utterances,
            audio_duration_ms=41_378,
            video_duration_ms=41_533,
            config=SegmentationConfig(),
            aligned_segments=aligned_segments,
        )

        self.assertIsNotNone(segment)
        assert segment is not None
        self.assertEqual(segment.end_ms, aligned_segments[0].start_ms - 1)


if __name__ == "__main__":
    unittest.main()
