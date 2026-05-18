from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from video_analysis_pipeline.exporter import export_review_page, export_workbook
from video_analysis_pipeline.models import Segment, SubtitleSpan
from video_analysis_pipeline.review_api import apply_review_adjustments


class ReviewApiTests(unittest.TestCase):
    def test_apply_review_adjustments_updates_generated_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_root = Path(tmp_dir) / "output"
            output_dir = output_root / "7"
            output_dir.mkdir(parents=True)

            segments = [
                Segment(
                    sequence_no=7,
                    segment_no=1,
                    text="Where are the other four?",
                    start_ms=18_613,
                    end_ms=21_033,
                    text_source="srt",
                    source_subtitle_index=0,
                    quality_flags=["subtitle_aligned"],
                ),
                Segment(
                    sequence_no=7,
                    segment_no=2,
                    text="I can't see them.",
                    start_ms=21_400,
                    end_ms=24_000,
                    text_source="srt",
                    source_subtitle_index=1,
                    quality_flags=["subtitle_aligned"],
                ),
            ]
            spans = [
                SubtitleSpan(
                    text="Where are the other four?",
                    normalized_text="where are the other four",
                    start_ms=18_613,
                    end_ms=21_033,
                    confidence=1.0,
                    frame_count=1,
                    source="srt",
                    raw_index=1,
                ),
                SubtitleSpan(
                    text="I can't see them.",
                    normalized_text="i can't see them",
                    start_ms=21_400,
                    end_ms=24_000,
                    confidence=1.0,
                    frame_count=1,
                    source="srt",
                    raw_index=2,
                ),
            ]

            manifest_path = output_dir / "manifest.json"
            segments_path = output_dir / "segments.json"
            subtitle_spans_path = output_dir / "subtitle_spans.json"
            segments_csv_path = output_dir / "segments.csv"
            review_path = output_dir / "review.html"
            workbook_path = output_dir / "dubbing.result.xlsx"
            batch_workbook_path = output_root / "dubbing.result.xlsx"
            source_mp4 = output_dir / "02.mp4"
            source_mp4.write_bytes(b"fake")

            manifest_path.write_text(
                json.dumps(
                    {
                        "sequence_no": 7,
                        "source_mp4": str(source_mp4),
                        "overview": {
                            "education_stage": "小学",
                            "subject": "[167070462398963715]英语",
                            "sequence_no": 7,
                            "movie_name": "Review Story",
                            "video_title": "Review Story",
                            "muted_video": "01.mp4",
                            "full_video": "02.mp4",
                            "background_audio": "03.mp3",
                            "cover_image": "01.jpg",
                            "video_description": "测试简介",
                            "difficulty": "",
                            "dialogue_audio": "",
                            "topic": "",
                            "source": "[7]绘本配音",
                        },
                        "outputs": {
                            "segments_csv": str(segments_csv_path),
                            "review_html": str(review_path),
                            "workbook": str(workbook_path),
                        },
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            segments_path.write_text(
                json.dumps(
                    {
                        "sequence_no": 7,
                        "subtitle": {
                            "spans": [item.to_json() for item in spans],
                        },
                        "segments": [item.to_json() for item in segments],
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            subtitle_spans_path.write_text(
                json.dumps(
                    {
                        "sequence_no": 7,
                        "subtitle_spans": [item.to_json() for item in spans],
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            export_review_page(review_path, video_path="02.mp4", segments=segments, title="Sequence 7 review")
            export_workbook(workbook_path, [item.to_excel_row() for item in segments])
            export_workbook(batch_workbook_path, [item.to_excel_row() for item in segments])
            (output_root / "batch_summary.json").write_text(
                json.dumps(
                    {
                        "items": [
                            {
                                "sequence_no": 7,
                                "output_dir": str(output_dir),
                                "segment_count": len(segments),
                            }
                        ]
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )

            result = apply_review_adjustments(
                review_page_path=review_path,
                adjustments=[
                    {
                        "segment_index": 0,
                        "segment_no": 1,
                        "start_ms": 18_500,
                        "end_ms": 20_800,
                    }
                ],
            )

            self.assertEqual(result["adjustment_count"], 1)
            self.assertTrue(str(segments_path) in result["updated_files"])
            self.assertTrue(str(batch_workbook_path) in result["updated_files"])

            updated_segments = json.loads(segments_path.read_text(encoding="utf-8"))
            self.assertEqual(updated_segments["segments"][0]["start_ms"], 18_500)
            self.assertEqual(updated_segments["segments"][0]["end_ms"], 20_800)
            self.assertEqual(updated_segments["subtitle"]["spans"][0]["start_ms"], 18_500)
            self.assertEqual(updated_segments["subtitle"]["spans"][0]["end_ms"], 20_800)

            updated_subtitle_spans = json.loads(subtitle_spans_path.read_text(encoding="utf-8"))
            self.assertEqual(updated_subtitle_spans["subtitle_spans"][0]["start_ms"], 18_500)
            self.assertEqual(updated_subtitle_spans["subtitle_spans"][0]["end_ms"], 20_800)

            csv_text = segments_csv_path.read_text(encoding="utf-8-sig")
            self.assertIn("00:00:18.500", csv_text)
            self.assertIn("00:00:20.800", csv_text)

            item_workbook = load_workbook(workbook_path)
            self.assertEqual(item_workbook.worksheets[0].cell(row=2, column=4).value, "Review Story")
            self.assertEqual(item_workbook.worksheets[1].cell(row=2, column=4).value, "00:00:18.500")
            self.assertEqual(item_workbook.worksheets[1].cell(row=2, column=5).value, "00:00:20.800")

            merged_workbook = load_workbook(batch_workbook_path)
            self.assertEqual(merged_workbook.worksheets[0].cell(row=2, column=4).value, "Review Story")
            self.assertEqual(merged_workbook.worksheets[1].cell(row=2, column=4).value, "00:00:18.500")
            self.assertEqual(merged_workbook.worksheets[1].cell(row=2, column=5).value, "00:00:20.800")


if __name__ == "__main__":
    unittest.main()
