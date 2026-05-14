from __future__ import annotations

import csv
import json
import html
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from video_analysis_pipeline.models import Segment


HEADERS = ["序号", "分视频序号", "分视频文本", "起始时间", "结束时间"]


def write_json(path: Path, payload: object) -> Path:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def export_csv(
    output_path: Path,
    rows: Iterable[tuple[int, int, str, str, str]],
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(rows)
    return output_path


def export_workbook(
    output_path: Path,
    rows: Iterable[tuple[int, int, str, str, str]],
    template_path: Path | None = None,
) -> Path:
    if template_path and template_path.exists():
        workbook = load_workbook(template_path)
    else:
        workbook = Workbook()

    while len(workbook.worksheets) < 2:
        workbook.create_sheet()

    worksheet = workbook.worksheets[1]

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    for column_index, header in enumerate(HEADERS, start=1):
        if worksheet.cell(row=1, column=column_index).value in (None, ""):
            worksheet.cell(row=1, column=column_index, value=header)

    row_index = 2
    for row in rows:
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)
        row_index += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def segments_to_rows(segments: list[Segment]) -> list[tuple[int, int, str, str, str]]:
    return [segment.to_excel_row() for segment in segments]


def export_review_page(
    output_path: Path,
    video_path: str,
    segments: list[Segment],
    title: str,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    segments_payload = json.dumps([segment.to_json() for segment in segments], ensure_ascii=False).replace("</", "<\\/")
    title_text = html.escape(title)
    video_src = html.escape(video_path)
    title_json = json.dumps(title, ensure_ascii=False).replace("</", "<\\/")
    video_path_json = json.dumps(video_path, ensure_ascii=False).replace("</", "<\\/")

    output_path.write_text(
        f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{title_text}</title>
  <style>
    :root {{
      color-scheme: light dark;
      font-family: Arial, sans-serif;
    }}
    body {{
      margin: 0;
      padding: 16px;
      background: #111827;
      color: #f9fafb;
    }}
    .layout {{
      display: grid;
      grid-template-columns: minmax(320px, 1.3fr) minmax(360px, 1fr);
      gap: 16px;
    }}
    .panel {{
      background: #1f2937;
      border-radius: 12px;
      padding: 16px;
      box-shadow: 0 8px 24px rgba(0, 0, 0, 0.2);
    }}
    video {{
      width: 100%;
      max-height: 70vh;
      border-radius: 8px;
      background: #000;
    }}
    .toolbar {{
      margin-top: 12px;
      display: flex;
      gap: 12px;
      align-items: center;
      flex-wrap: wrap;
      font-size: 14px;
    }}
    .segment-list {{
      display: grid;
      gap: 8px;
      max-height: 78vh;
      overflow: auto;
    }}
    .filter-row {{
      display: flex;
      align-items: center;
      gap: 8px;
      margin-bottom: 12px;
      font-size: 14px;
    }}
    .editor-panel {{
      margin-top: 16px;
      border: 1px solid #374151;
      border-radius: 10px;
      padding: 12px;
      background: #0f172a;
    }}
    .editor-title {{
      margin: 0 0 8px;
      font-size: 15px;
      font-weight: 600;
    }}
    .editor-summary {{
      font-size: 13px;
      line-height: 1.45;
      color: #cbd5e1;
      margin-bottom: 12px;
      white-space: pre-wrap;
    }}
    .edit-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 10px 12px;
      margin-bottom: 12px;
    }}
    .edit-field {{
      display: grid;
      gap: 6px;
      font-size: 13px;
    }}
    .edit-field input {{
      width: 100%;
      box-sizing: border-box;
      border: 1px solid #475569;
      border-radius: 6px;
      padding: 8px 10px;
      background: #111827;
      color: #f9fafb;
    }}
    .edit-actions {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-bottom: 10px;
    }}
    .save-status {{
      font-size: 12px;
      color: #93c5fd;
      min-height: 18px;
    }}
    .segment-item {{
      border: 1px solid #374151;
      border-radius: 8px;
      padding: 10px 12px;
      cursor: pointer;
      background: #111827;
    }}
    .segment-item:hover {{
      border-color: #60a5fa;
    }}
    .segment-item.active {{
      border-color: #22c55e;
      background: #0f2d22;
    }}
    .segment-item.adjusted {{
      border-color: #f59e0b;
    }}
    .segment-meta {{
      color: #93c5fd;
      font-size: 12px;
      margin-bottom: 4px;
    }}
    .segment-text {{
      font-size: 15px;
      line-height: 1.45;
      white-space: pre-wrap;
    }}
    .segment-flags {{
      margin-top: 6px;
      color: #fca5a5;
      font-size: 12px;
    }}
    .segment-adjustment {{
      margin-top: 6px;
      color: #fbbf24;
      font-size: 12px;
    }}
    button {{
      border: 0;
      border-radius: 6px;
      padding: 8px 12px;
      background: #2563eb;
      color: white;
      cursor: pointer;
    }}
    button.secondary {{
      background: #374151;
    }}
    button:disabled {{
      opacity: 0.55;
      cursor: not-allowed;
    }}
    @media (max-width: 980px) {{
      .layout {{
        grid-template-columns: 1fr;
      }}
      .segment-list {{
        max-height: none;
      }}
    }}
  </style>
</head>
<body>
  <div class="layout">
    <section class="panel">
      <h1 style="margin-top: 0;">配音时间校验页</h1>
      <video id="video" controls preload="metadata" src="{video_src}"></video>
      <div class="toolbar">
        <span id="status">点击右侧分段可跳转并播放。</span>
        <button id="replayButton" class="secondary" type="button">重播当前段</button>
      </div>
      <div class="editor-panel">
        <h2 class="editor-title">手工微调</h2>
        <div id="editorSummary" class="editor-summary">先在右侧选择一个分段，再按毫秒步长微调开始/结束时间。</div>
        <div class="edit-grid">
          <label class="edit-field">
            <span>向前步长（毫秒）</span>
            <input id="backwardStepMs" type="number" min="1" step="1" value="15">
          </label>
          <label class="edit-field">
            <span>向后步长（毫秒）</span>
            <input id="forwardStepMs" type="number" min="1" step="1" value="15">
          </label>
        </div>
        <div class="edit-actions">
          <button id="moveStartBackwardButton" type="button">开始时间 - 向前</button>
          <button id="moveStartForwardButton" class="secondary" type="button">开始时间 + 向后</button>
          <button id="moveEndBackwardButton" class="secondary" type="button">结束时间 - 向前</button>
          <button id="moveEndForwardButton" type="button">结束时间 + 向后</button>
        </div>
        <div class="edit-actions">
          <button id="resetSegmentButton" class="secondary" type="button">重置当前段</button>
          <button id="saveAdjustmentsButton" type="button">保存并回写文件</button>
        </div>
        <div id="saveStatus" class="save-status">未保存手工修正。启动 review-server 后可直接回写输出文件。</div>
      </div>
    </section>
    <section class="panel">
      <h2 style="margin-top: 0;">分段列表</h2>
      <label class="filter-row">
        <input id="onlyAnomalies" type="checkbox">
        <span>只看异常段</span>
      </label>
      <div id="segmentList" class="segment-list"></div>
    </section>
  </div>
  <script id="segments-data" type="application/json">{segments_payload}</script>
  <script>
    const originalSegments = JSON.parse(document.getElementById('segments-data').textContent);
    const segments = originalSegments.map(segment => ({{ ...segment }}));
    const benignFlags = new Set(['subtitle_aligned']);
    const reviewTitle = {title_json};
    const reviewVideoPath = {video_path_json};
    const defaultApiBase = 'http://127.0.0.1:8765';
    const storageKey = `video-analysis-review-adjustments::${{window.location.pathname}}`;
    const video = document.getElementById('video');
    const segmentList = document.getElementById('segmentList');
    const status = document.getElementById('status');
    const onlyAnomalies = document.getElementById('onlyAnomalies');
    const replayButton = document.getElementById('replayButton');
    const backwardStepInput = document.getElementById('backwardStepMs');
    const forwardStepInput = document.getElementById('forwardStepMs');
    const moveStartBackwardButton = document.getElementById('moveStartBackwardButton');
    const moveStartForwardButton = document.getElementById('moveStartForwardButton');
    const moveEndBackwardButton = document.getElementById('moveEndBackwardButton');
    const moveEndForwardButton = document.getElementById('moveEndForwardButton');
    const resetSegmentButton = document.getElementById('resetSegmentButton');
    const saveAdjustmentsButton = document.getElementById('saveAdjustmentsButton');
    const editorSummary = document.getElementById('editorSummary');
    const saveStatus = document.getElementById('saveStatus');
    let activeSegmentIndex = -1;
    let playbackEndSeconds = null;

    function formatTimecode(milliseconds) {{
      const total = Math.max(0, Math.trunc(milliseconds));
      const hours = Math.floor(total / 3600000);
      const minutes = Math.floor((total % 3600000) / 60000);
      const seconds = Math.floor((total % 60000) / 1000);
      const ms = total % 1000;
      return [
        String(hours).padStart(2, '0'),
        String(minutes).padStart(2, '0'),
        String(seconds).padStart(2, '0'),
      ].join(':') + `.${{String(ms).padStart(3, '0')}}`;
    }}

    function formatRange(segment) {{
      return `${{segment.segment_no}} | ${{formatTimecode(segment.start_ms)}} - ${{formatTimecode(segment.end_ms)}}`;
    }}

    function escapeHtml(text) {{
      return text
        .replaceAll('&', '&amp;')
        .replaceAll('<', '&lt;')
        .replaceAll('>', '&gt;')
        .replaceAll('"', '&quot;')
        .replaceAll("'", '&#39;');
    }}

    function getAnomalyFlags(segment) {{
      if (!Array.isArray(segment.quality_flags)) {{
        return [];
      }}
      return segment.quality_flags.filter(flag => !benignFlags.has(flag));
    }}

    function isAdjusted(index) {{
      const original = originalSegments[index];
      const current = segments[index];
      return Boolean(original && current) && (
        original.start_ms !== current.start_ms || original.end_ms !== current.end_ms
      );
    }}

    function formatSignedMilliseconds(delta) {{
      const value = Math.trunc(delta);
      return `${{value >= 0 ? '+' : ''}}${{value}}ms`;
    }}

    function getStepValue(input) {{
      const parsed = Number.parseInt(input.value, 10);
      return Number.isFinite(parsed) && parsed > 0 ? parsed : 15;
    }}

    function getMinimumStart(index) {{
      if (index <= 0) {{
        return 0;
      }}
      return segments[index - 1].end_ms + 1;
    }}

    function getMaximumEnd(index) {{
      if (index >= segments.length - 1) {{
        if (Number.isFinite(video.duration) && video.duration > 0) {{
          return Math.round(video.duration * 1000);
        }}
        return Number.MAX_SAFE_INTEGER;
      }}
      return segments[index + 1].start_ms - 1;
    }}

    function normalizeSegmentTiming(index) {{
      const segment = segments[index];
      if (!segment) {{
        return;
      }}
      const minimumStart = getMinimumStart(index);
      const maximumEnd = getMaximumEnd(index);
      segment.start_ms = Math.max(0, Math.max(minimumStart, Math.trunc(segment.start_ms)));
      segment.end_ms = Math.min(maximumEnd, Math.trunc(segment.end_ms));
      if (segment.end_ms <= segment.start_ms) {{
        segment.end_ms = Math.min(maximumEnd, segment.start_ms + 1);
      }}
      if (segment.end_ms <= segment.start_ms) {{
        segment.start_ms = Math.max(0, Math.min(segment.start_ms, segment.end_ms - 1));
      }}
      segment.duration_ms = Math.max(0, segment.end_ms - segment.start_ms);
      segment.start_timecode = formatTimecode(segment.start_ms);
      segment.end_timecode = formatTimecode(segment.end_ms);
    }}

    function updateEditorSummary() {{
      if (activeSegmentIndex < 0 || !segments[activeSegmentIndex]) {{
        editorSummary.textContent = '先在右侧选择一个分段，再按毫秒步长微调开始/结束时间。';
        return;
      }}
      const segment = segments[activeSegmentIndex];
      const original = originalSegments[activeSegmentIndex];
      const startDelta = segment.start_ms - original.start_ms;
      const endDelta = segment.end_ms - original.end_ms;
      editorSummary.textContent = [
        `当前分段：${{segment.segment_no}} / ${{segment.text}}`,
        `当前时间：${{formatTimecode(segment.start_ms)}} - ${{formatTimecode(segment.end_ms)}}`,
        `原始时间：${{formatTimecode(original.start_ms)}} - ${{formatTimecode(original.end_ms)}}`,
        `修正量：开始 ${{formatSignedMilliseconds(startDelta)}}，结束 ${{formatSignedMilliseconds(endDelta)}}`,
      ].join('\\n');
    }}

    function collectAdjustments() {{
      return segments
        .map((segment, index) => {{
          if (!isAdjusted(index)) {{
            return null;
          }}
          const original = originalSegments[index];
          return {{
            segment_index: index,
            segment_no: segment.segment_no,
            text: segment.text,
            original_start_ms: original.start_ms,
            original_end_ms: original.end_ms,
            start_ms: segment.start_ms,
            end_ms: segment.end_ms,
            delta_start_ms: segment.start_ms - original.start_ms,
            delta_end_ms: segment.end_ms - original.end_ms,
            quality_flags: segment.quality_flags || [],
          }};
        }})
        .filter(Boolean);
    }}

    function buildAdjustmentsPayload() {{
      return {{
        version: 2,
        saved_at: new Date().toISOString(),
        review_title: reviewTitle,
        review_path: window.location.pathname,
        video_path: reviewVideoPath,
        segment_count: segments.length,
        adjustments: collectAdjustments(),
      }};
    }}

    function getReviewApiBase() {{
      if (window.location.protocol === 'http:' || window.location.protocol === 'https:') {{
        return window.location.origin;
      }}
      return defaultApiBase;
    }}

    function cacheAdjustments(payload) {{
      localStorage.setItem(storageKey, JSON.stringify(payload));
    }}

    function commitSavedAdjustments() {{
      segments.forEach((segment, index) => {{
        Object.assign(originalSegments[index], {{
          start_ms: segment.start_ms,
          end_ms: segment.end_ms,
          duration_ms: segment.duration_ms,
          start_timecode: segment.start_timecode,
          end_timecode: segment.end_timecode,
        }});
      }});
      localStorage.removeItem(storageKey);
      renderSegments();
    }}

    async function postAdjustmentsToApi(payload) {{
      const response = await fetch(`${{getReviewApiBase()}}/api/review/save`, {{
        method: 'POST',
        headers: {{
          'Content-Type': 'application/json',
        }},
        body: JSON.stringify(payload),
      }});
      const result = await response.json().catch(() => ({{ ok: false, error: `HTTP ${{response.status}}` }}));
      if (!response.ok || !result.ok) {{
        throw new Error(result.error || `HTTP ${{response.status}}`);
      }}
      return result;
    }}

    async function saveAdjustments() {{
      const payload = buildAdjustmentsPayload();
      if (payload.adjustments.length === 0) {{
        saveStatus.textContent = '当前没有新的时间修正可保存。';
        return null;
      }}
      cacheAdjustments(payload);
      saveStatus.textContent = '正在回写 segments.json / CSV / Excel ...';
      try {{
        const result = await postAdjustmentsToApi(payload);
        commitSavedAdjustments();
        saveStatus.textContent = `已回写 ${{result.adjustment_count}} 条修正，并同步更新正式输出文件。`;
        status.textContent = `保存完成：已同步更新 ${{result.updated_files.length}} 个文件。`;
        return result;
      }} catch (error) {{
        console.error(error);
        saveStatus.textContent = `回写失败：${{error.message}}。请先运行 python -m video_analysis_pipeline review-server --output-root output`; 
        return null;
      }}
    }}

    function applySavedAdjustments() {{
      const saved = localStorage.getItem(storageKey);
      if (!saved) {{
        return;
      }}
      try {{
        const payload = JSON.parse(saved);
        const savedAdjustments = Array.isArray(payload.adjustments) ? payload.adjustments : [];
        savedAdjustments
          .slice()
          .sort((left, right) => left.segment_no - right.segment_no)
          .forEach((adjustment) => {{
            const index = segments.findIndex(segment => segment.segment_no === adjustment.segment_no);
            if (index < 0) {{
              return;
            }}
            segments[index].start_ms = adjustment.start_ms;
            segments[index].end_ms = adjustment.end_ms;
            normalizeSegmentTiming(index);
          }});
        if (savedAdjustments.length > 0) {{
          saveStatus.textContent = `已恢复 ${{savedAdjustments.length}} 条未落盘修正。`;
        }}
      }} catch (error) {{
        console.error(error);
        saveStatus.textContent = '读取已保存修正失败，已忽略旧数据。';
      }}
    }}

    function updateActionButtons() {{
      const disabled = activeSegmentIndex < 0;
      moveStartBackwardButton.disabled = disabled;
      moveStartForwardButton.disabled = disabled;
      moveEndBackwardButton.disabled = disabled;
      moveEndForwardButton.disabled = disabled;
      resetSegmentButton.disabled = disabled;
    }}

    function setActiveSegment(index) {{
      activeSegmentIndex = index;
      [...segmentList.children].forEach((item) => {{
        item.classList.toggle('active', Number(item.dataset.index) === index);
      }});
      if (index >= 0) {{
        [...segmentList.children]
          .find(item => Number(item.dataset.index) === index)
          ?.scrollIntoView({{ block: 'nearest' }});
      }}
      updateEditorSummary();
      updateActionButtons();
    }}

    function renderSegments() {{
      segmentList.innerHTML = '';
      segments.forEach((segment, index) => {{
        const anomalyFlags = getAnomalyFlags(segment);
        if (onlyAnomalies.checked && anomalyFlags.length === 0) {{
          return;
        }}
        const item = document.createElement('div');
        item.className = 'segment-item';
        if (isAdjusted(index)) {{
          item.classList.add('adjusted');
        }}
        item.dataset.index = String(index);
        const adjustmentSummary = isAdjusted(index)
          ? `<div class="segment-adjustment">手工修正：开始 ${{formatSignedMilliseconds(segments[index].start_ms - originalSegments[index].start_ms)}} / 结束 ${{formatSignedMilliseconds(segments[index].end_ms - originalSegments[index].end_ms)}}</div>`
          : '';
        item.innerHTML = `
          <div class="segment-meta">${{formatRange(segment)}}</div>
          <div class="segment-text">${{escapeHtml(segment.text)}}</div>
          <div class="segment-flags">${{(segment.quality_flags || []).join(' | ')}}</div>
          ${{adjustmentSummary}}
        `;
        item.addEventListener('click', () => playSegment(index));
        segmentList.appendChild(item);
      }});
      setActiveSegment(activeSegmentIndex);
    }}

    function playSegment(index) {{
      const segment = segments[index];
      if (!segment) {{
        return;
      }}
      playbackEndSeconds = segment.end_ms / 1000;
      video.currentTime = segment.start_ms / 1000;
      setActiveSegment(index);
      status.textContent = `当前分段：${{segment.segment_no}} / ${{segment.text}}`;
      video.play().catch(() => {{
        status.textContent = '浏览器阻止了自动播放，请手动点击播放。';
      }});
    }}

    function refreshActivePlaybackWindow() {{
      if (activeSegmentIndex < 0) {{
        return;
      }}
      const segment = segments[activeSegmentIndex];
      if (!segment) {{
        return;
      }}
      playbackEndSeconds = segment.end_ms / 1000;
      const currentMs = video.currentTime * 1000;
      if (currentMs < segment.start_ms || currentMs > segment.end_ms) {{
        video.currentTime = segment.start_ms / 1000;
      }}
    }}

    function describeBoundary(boundary) {{
      return boundary === 'start_ms' ? '开始时间' : '结束时间';
    }}

    function describeDirection(direction) {{
      return direction < 0 ? '向前' : '向后';
    }}

    function adjustSegmentBoundary(boundary, direction) {{
      if (activeSegmentIndex < 0) {{
        return;
      }}
      const step = getStepValue(direction < 0 ? backwardStepInput : forwardStepInput);
      const segment = segments[activeSegmentIndex];
      const previousValue = segment[boundary];
      segment[boundary] += direction * step;
      normalizeSegmentTiming(activeSegmentIndex);
      const actualDelta = segment[boundary] - previousValue;
      renderSegments();
      refreshActivePlaybackWindow();
      if (actualDelta === 0) {{
        status.textContent = `${{describeBoundary(boundary)}}已到可调整边界。`;
        return;
      }}
      status.textContent = `已将分段 ${{segment.segment_no}} 的${{describeBoundary(boundary)}}${{describeDirection(direction)}}调整 ${{Math.abs(actualDelta)}}ms。`;
    }}

    function resetCurrentSegment() {{
      if (activeSegmentIndex < 0) {{
        return;
      }}
      const original = originalSegments[activeSegmentIndex];
      segments[activeSegmentIndex].start_ms = original.start_ms;
      segments[activeSegmentIndex].end_ms = original.end_ms;
      normalizeSegmentTiming(activeSegmentIndex);
      renderSegments();
      refreshActivePlaybackWindow();
      status.textContent = `已重置分段 ${{segments[activeSegmentIndex].segment_no}} 的手工修正。`;
    }}

    replayButton.addEventListener('click', () => {{
      if (activeSegmentIndex >= 0) {{
        playSegment(activeSegmentIndex);
      }}
    }});
    moveStartBackwardButton.addEventListener('click', () => adjustSegmentBoundary('start_ms', -1));
    moveStartForwardButton.addEventListener('click', () => adjustSegmentBoundary('start_ms', 1));
    moveEndBackwardButton.addEventListener('click', () => adjustSegmentBoundary('end_ms', -1));
    moveEndForwardButton.addEventListener('click', () => adjustSegmentBoundary('end_ms', 1));
    resetSegmentButton.addEventListener('click', resetCurrentSegment);
    saveAdjustmentsButton.addEventListener('click', () => {{
      void saveAdjustments();
    }});
    onlyAnomalies.addEventListener('change', () => renderSegments());

    video.addEventListener('timeupdate', () => {{
      if (playbackEndSeconds !== null && video.currentTime >= playbackEndSeconds) {{
        video.pause();
      }}
      const currentMs = video.currentTime * 1000;
      const index = segments.findIndex(segment => currentMs >= segment.start_ms && currentMs <= segment.end_ms);
      if (index >= 0 && index !== activeSegmentIndex) {{
        setActiveSegment(index);
      }}
    }});

    applySavedAdjustments();
    updateActionButtons();
    renderSegments();
  </script>
</body>
</html>
""",
        encoding="utf-8",
    )
    return output_path
